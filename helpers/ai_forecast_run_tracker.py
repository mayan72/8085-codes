import os
import re
import tempfile
from datetime import datetime, timedelta
from zipfile import ZipFile, is_zipfile

from django.conf import settings
from django.utils import timezone
from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from helpers.config import base_forecast_logger as logger

TRACKING_SUBDIR = "ai_forecast_tracking"
TRACKING_FILENAME = "ai_forecast_cp_tracking.xlsx"
RETENTION_DAYS = 2

HEADERS = [
    "run_at",
    "job_id",
    "cp_id",
    "cp_name",
    "status",
    "format_ok",
    "correctness_C",
    "expected_direction",
    "detected_direction",
    "forecast_status",
    "summary_status",
    "forecast_input_tokens",
    "forecast_output_tokens",
    "forecast_reasoning_tokens",
    "forecast_total_tokens",
    "summary_input_tokens",
    "summary_output_tokens",
    "summary_reasoning_tokens",
    "summary_total_tokens",
    "total_tokens",
    "news_dump_count",
    "price_url_count",
    "rows_saved",
    "errors",
    "warnings",
]

WRAP_COLUMNS = ("errors", "warnings")

RISE_RE = re.compile(
    r"\b(rise|rises|rising|increase|increased|increasing|higher|upward)\b",
    re.I,
)
FALL_RE = re.compile(
    r"\b(fall|falls|falling|decrease|decreased|decreasing|decline|declining|lower|downward|drop|dropped)\b",
    re.I,
)
STABLE_RE = re.compile(
    r"\b(stable|unchanged|flat|sideways|steady)\b",
    re.I,
)


class TrackingExtractError(ValueError):
    """Raised when the first Short-term Outlook bullet cannot be parsed."""


def tracking_excel_path():
    media_root = getattr(settings, "MEDIA_ROOT", None)
    if not media_root:
        raise ValueError("MEDIA_ROOT is not configured")
    folder = os.path.join(str(media_root), TRACKING_SUBDIR)
    os.makedirs(folder, exist_ok=True)
    return os.path.join(folder, TRACKING_FILENAME)


def _extract_first_short_term_bullet(analysis):
    """Return (original_bullet, start_index, end_index) for the first Short-term Ω bullet only."""
    text = analysis or ""

    short_match = re.search(
        r"(?im)^\s*\*{0,2}Short-term Outlook\*{0,2}\s*:?",
        text,
    )
    if not short_match:
        raise TrackingExtractError("BASE_ANALYSIS is missing 'Short-term Outlook'")

    section_start = short_match.end()
    long_match = re.search(
        r"(?im)^\s*\*{0,2}Long-term Outlook\*{0,2}\s*:?",
        text[section_start:],
    )
    section_end = section_start + long_match.start() if long_match else len(text)
    section = text[section_start:section_end]

    omega_match = re.search(r"Ω[^\r\n]*", section)
    if omega_match:
        bullet_start = section_start + omega_match.start()
        bullet_end = section_start + omega_match.end()
        original = text[bullet_start:bullet_end].strip()
        if not original:
            raise TrackingExtractError("Short-term Outlook first bullet is empty")
        return original, bullet_start, bullet_end

    bullet_start = section_start
    while bullet_start < section_end and text[bullet_start].isspace():
        bullet_start += 1

    line_end = text.find("\n", bullet_start, section_end)
    bullet_end = line_end if line_end != -1 else section_end
    while bullet_end > bullet_start and text[bullet_end - 1].isspace():
        bullet_end -= 1

    original = text[bullet_start:bullet_end].strip()
    if not original:
        raise TrackingExtractError("Short-term Outlook section is empty")
    return original, bullet_start, bullet_end


def usage_from_response(response):
    empty = {
        "input_tokens": None,
        "output_tokens": None,
        "reasoning_tokens": None,
        "total_tokens": None,
    }
    if not response:
        return empty
    try:
        usage = getattr(response, "usage", None)
        if usage is None:
            return empty
        if isinstance(usage, dict):
            output_details = usage.get("output_tokens_details") or {}
            reasoning = (
                output_details.get("reasoning_tokens")
                if isinstance(output_details, dict)
                else None
            )
            return {
                "input_tokens": usage.get("input_tokens"),
                "output_tokens": usage.get("output_tokens"),
                "reasoning_tokens": reasoning,
                "total_tokens": usage.get("total_tokens"),
            }
        output_details = getattr(usage, "output_tokens_details", None)
        reasoning = None
        if isinstance(output_details, dict):
            reasoning = output_details.get("reasoning_tokens")
        elif output_details is not None:
            reasoning = getattr(output_details, "reasoning_tokens", None)
        return {
            "input_tokens": getattr(usage, "input_tokens", None),
            "output_tokens": getattr(usage, "output_tokens", None),
            "reasoning_tokens": reasoning,
            "total_tokens": getattr(usage, "total_tokens", None),
        }
    except Exception as exc:
        logger.exception("[TRACKER] Failed to read OpenAI usage: %s", exc)
        return empty


def required_price_direction(last_actual, new_forecast):
    try:
        last_val = float(last_actual)
        new_val = float(new_forecast)
    except (TypeError, ValueError):
        return None
    if new_val > last_val:
        return "rise"
    if new_val < last_val:
        return "fall"
    return "stable"


def detect_bullet_direction(text):
    text = text or ""
    has_rise = bool(RISE_RE.search(text))
    has_fall = bool(FALL_RE.search(text))
    has_stable = bool(STABLE_RE.search(text))
    if has_rise and not has_fall:
        return "rise"
    if has_fall and not has_rise:
        return "fall"
    if has_stable and not has_rise and not has_fall:
        return "stable"
    if has_rise and has_fall:
        return "mixed"
    return "unclear"


def score_correctness_c(last_actual, new_forecast, replacement_bullet):
    try:
        expected = required_price_direction(last_actual, new_forecast)
        if expected is None:
            return "N/A", None, "unclear", "last_actual or new_forecast missing/non-numeric"
        detected = detect_bullet_direction(replacement_bullet)
        if detected == expected:
            return "PASS", expected, detected, None
        if detected == "mixed" and expected in {"rise", "fall"}:
            return "PASS", expected, detected, "bullet mentions both rise and fall"
        if detected == "unclear":
            return "UNCLEAR", expected, detected, "no clear rise/fall/stable wording"
        return "FAIL", expected, detected, "summary direction does not match last actual vs new forecast"
    except Exception as exc:
        logger.exception("[TRACKER] score_correctness_c failed: %s", exc)
        return "UNCLEAR", None, "unclear", str(exc)


def score_faithfulness_f(original_analysis, merged_analysis, bullet_start, bullet_end):
    try:
        if not original_analysis or not merged_analysis:
            return "FAIL", "missing original or merged analysis"
        if bullet_start is None or bullet_end is None:
            return "FAIL", "missing original bullet span"
        original_rest = original_analysis[:bullet_start] + original_analysis[bullet_end:]
        _, merged_start, merged_end = _extract_first_short_term_bullet(merged_analysis)
        merged_rest = merged_analysis[:merged_start] + merged_analysis[merged_end:]
        if original_rest.strip() == merged_rest.strip():
            return "PASS", None
        return "FAIL", "text outside first short-term bullet changed"
    except TrackingExtractError as exc:
        return "FAIL", str(exc)
    except Exception as exc:
        logger.exception("[TRACKER] score_faithfulness_f failed: %s", exc)
        return "FAIL", str(exc)


def score_faithfulness_g(expected_direction, replacement_bullet):
    try:
        if not expected_direction:
            return "N/A", "no expected direction"
        detected = detect_bullet_direction(replacement_bullet)
        if detected == expected_direction:
            return "PASS", None
        if detected == "mixed" and expected_direction in {"rise", "fall"}:
            return "PASS", "two-month wording present; required direction also present"
        if detected == "unclear":
            return "UNCLEAR", "could not detect direction wording"
        if expected_direction == "rise" and detected == "fall":
            return "FAIL", "replacement contradicts required rise"
        if expected_direction == "fall" and detected == "rise":
            return "FAIL", "replacement contradicts required fall"
        if expected_direction == "stable" and detected in {"rise", "fall"}:
            return "FAIL", "replacement contradicts required stable"
        return "UNCLEAR", "detected=%s" % detected
    except Exception as exc:
        logger.exception("[TRACKER] score_faithfulness_g failed: %s", exc)
        return "UNCLEAR", str(exc)


def _nonempty_url(value):
    return bool(str(value or "").strip())


def score_faithfulness_h(forecast_payload):
    try:
        if not isinstance(forecast_payload, dict):
            return "N/A", "no forecast payload", 0, 0

        price_data = forecast_payload.get("price_data") or {}
        news_dump = forecast_payload.get("news_dump") or []
        news_used = forecast_payload.get("news_inputs_used") or {}

        price_claimed = bool(price_data.get("reliable_price_data"))
        try:
            news_count_used = int(news_used.get("news_count_used") or 0)
        except (TypeError, ValueError):
            news_count_used = 0
        news_claimed = news_count_used > 0

        url_count = 0
        for bucket in (
            price_data.get("daily_or_weekly_prices_current_month") or [],
            price_data.get("daily_or_weekly_prices_last_month") or [],
            price_data.get("sources") or [],
        ):
            if not isinstance(bucket, list):
                continue
            for item in bucket:
                if isinstance(item, dict) and _nonempty_url(item.get("url")):
                    url_count += 1

        news_url_count = 0
        if isinstance(news_dump, list):
            for item in news_dump:
                if isinstance(item, dict) and _nonempty_url(item.get("url")):
                    news_url_count += 1

        if not price_claimed and not news_claimed:
            return "N/A", None, news_url_count, url_count

        notes = []
        failed = False
        if price_claimed and url_count == 0:
            failed = True
            notes.append("price data claimed reliable but no price/source URLs")
        if news_claimed and news_url_count == 0:
            failed = True
            notes.append("news claimed used but news_dump has no URLs")

        if failed:
            return "FAIL", "; ".join(notes), news_url_count, url_count
        return "PASS", None, news_url_count, url_count
    except Exception as exc:
        logger.exception("[TRACKER] score_faithfulness_h failed: %s", exc)
        return "UNCLEAR", str(exc), 0, 0


def _parse_run_at(run_at):
    if run_at is None or run_at == "":
        return None
    if isinstance(run_at, datetime):
        parsed = run_at
    else:
        try:
            parsed = datetime.strptime(str(run_at).strip(), "%Y-%m-%d %H:%M:%S")
        except (TypeError, ValueError):
            return None
    if timezone.is_naive(parsed):
        return timezone.make_aware(parsed, timezone.get_current_timezone())
    return parsed


def _safe_excel_value(value):
    """Coerce values to Excel-safe scalars that LibreOffice and Excel can both open."""
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, (int, float)):
        return value
    text = str(value)
    text = ILLEGAL_CHARACTERS_RE.sub("", text)
    # Excel/LibreOffice cell limit is 32767 characters.
    if len(text) > 32767:
        text = text[:32767]
    return text


def _load_existing_rows(path, cutoff):
    existing = []

    if not os.path.exists(path):
        return existing

    if not is_zipfile(path):
        logger.warning(
            "[TRACKER] Existing tracking file is not a valid xlsx zip: %s. Starting fresh.",
            path,
        )
        return existing

    workbook = None

    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet = workbook.active

        if sheet.max_row and sheet.max_row > 10000:
            logger.warning(
                "[TRACKER] Existing tracking Excel has unexpectedly many rows: %s. "
                "Ignoring existing rows and creating a fresh file.",
                sheet.max_row,
            )
            return existing

        rows = sheet.iter_rows(
            min_row=1,
            max_row=min(sheet.max_row or 1, 1048576),
            values_only=True,
        )

        header = next(rows, None)
        if not header:
            return existing

        header = [str(col).strip() if col is not None else "" for col in header]

        for values in rows:
            record = dict(zip(header, values))
            parsed = _parse_run_at(record.get("run_at"))
            if parsed is None or parsed >= cutoff:
                existing.append(record)

        return existing

    except Exception as exc:
        logger.exception(
            "[TRACKER] Failed to read existing Excel %s: %s",
            path,
            exc,
        )
        return existing

    finally:
        if workbook is not None:
            workbook.close()


def _apply_sheet_layout(sheet):
    wrap = Alignment(wrap_text=True, vertical="top")
    header_font = Font(bold=True)

    widths = {
        "run_at": 20,
        "job_id": 18,
        "cp_id": 14,
        "cp_name": 28,
        "status": 12,
        "errors": 60,
        "warnings": 60,
    }

    wrap_indexes = []
    for index, name in enumerate(HEADERS, start=1):
        letter = get_column_letter(index)
        sheet.column_dimensions[letter].width = widths.get(name, 18)
        sheet.cell(row=1, column=index).font = header_font
        if name in WRAP_COLUMNS:
            wrap_indexes.append(index)

    max_row = min(sheet.max_row or 1, 1048576)
    for row_number in range(2, max_row + 1):
        for column_index in wrap_indexes:
            sheet.cell(row=row_number, column=column_index).alignment = wrap


def _atomic_save_workbook(workbook, path):
    directory = os.path.dirname(path) or "."
    fd, tmp_path = tempfile.mkstemp(suffix=".xlsx", dir=directory)
    os.close(fd)
    try:
        workbook.save(tmp_path)
        workbook.close()
        if not is_zipfile(tmp_path):
            raise ValueError("Saved workbook is not a valid xlsx zip archive")
        with ZipFile(tmp_path, "r") as archive:
            required = {
                "[Content_Types].xml",
                "xl/workbook.xml",
                "xl/worksheets/sheet1.xml",
            }
            missing = required - set(archive.namelist())
            if missing:
                raise ValueError("Saved workbook is missing required parts: %s" % sorted(missing))
        os.replace(tmp_path, path)
    except Exception:
        try:
            os.remove(tmp_path)
        except OSError:
            pass
        raise


class ForecastRunTracker:
    def __init__(self):
        self.job_id = timezone.now().strftime("%Y%m%dT%H%M%S")
        self.rows = []

    def add_row(self, **kwargs):
        try:
            row = {key: kwargs.get(key, "") for key in HEADERS}
            row["run_at"] = timezone.localtime().strftime("%Y-%m-%d %H:%M:%S")
            row["job_id"] = self.job_id
            self.rows.append(row)
        except Exception as exc:
            logger.exception("[TRACKER] Failed to add tracking row: %s", exc)

    def flush(self):
        if not self.rows:
            return None
        workbook = None
        try:
            path = tracking_excel_path()
            cutoff = timezone.localtime() - timedelta(days=RETENTION_DAYS)
            existing = _load_existing_rows(path, cutoff)

            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "cp_tracking"
            sheet.append(HEADERS)
            for record in existing + self.rows:
                sheet.append([_safe_excel_value(record.get(key, "")) for key in HEADERS])
            _apply_sheet_layout(sheet)
            _atomic_save_workbook(workbook, path)
            workbook = None
            logger.info("Wrote %s tracking rows to %s", len(self.rows), path)
            return path
        except Exception as exc:
            logger.exception("[TRACKER] Failed to write tracking Excel: %s", exc)
            return None
        finally:
            if workbook is not None:
                try:
                    workbook.close()
                except Exception:
                    pass
